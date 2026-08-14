from pathlib import Path
import sys
from openpyxl import load_workbook


WORKBOOK = Path(r"C:\Users\cwbec\BlockMed\financial-projections\Blockmediary_Financial_Model_Submission_Ready_2026-08-13.xlsx")
sys.stdout.reconfigure(encoding="utf-8")


def display(value):
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.8g}"
    text = str(value).replace("\n", " ").strip()
    return text[:180]


wb_formula = load_workbook(WORKBOOK, data_only=False, read_only=False)
wb_value = load_workbook(WORKBOOK, data_only=True, read_only=False)

print("SHEETS")
for ws in wb_formula.worksheets:
    formulas = 0
    nonempty = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                nonempty += 1
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas += 1
    print(f"- {ws.title}: {ws.max_row} rows x {ws.max_column} cols; {nonempty} populated; {formulas} formulas")

targets = {
    "Dashboard": range(1, 85),
    "Launch Readiness": range(1, 45),
    "Pitch Deck References": range(1, 75),
    "Model Checks": range(1, 85),
    "P&L 5yr": range(1, 70),
    "Cash Flow Statement": range(1, 60),
    "Tier Economics": range(1, 45),
    "Client Funds Memo": range(1, 30),
    "Scenario Engine": range(1, 70),
}

for sheet_name, rows in targets.items():
    if sheet_name not in wb_formula.sheetnames:
        continue
    ws_f = wb_formula[sheet_name]
    ws_v = wb_value[sheet_name]
    print(f"\n[{sheet_name}]")
    for row_idx in rows:
        values = []
        has_content = False
        for col_idx in range(1, min(ws_f.max_column, 18) + 1):
            formula = ws_f.cell(row_idx, col_idx).value
            cached = ws_v.cell(row_idx, col_idx).value
            if formula is not None or cached is not None:
                has_content = True
            if isinstance(formula, str) and formula.startswith("="):
                value = f"{display(cached)} <{display(formula)}>"
            else:
                value = display(cached if cached is not None else formula)
            values.append(value)
        if has_content:
            while values and not values[-1]:
                values.pop()
            print(f"{row_idx}: " + " | ".join(values))
