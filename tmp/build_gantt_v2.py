from copy import copy
from datetime import datetime, timedelta
from pathlib import Path
from tempfile import NamedTemporaryFile
from zipfile import ZipFile, ZIP_DEFLATED
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
V3 = ROOT / "Project Governance" / "v3"
SOURCE = V3 / "master-gantt.xlsx"
OUTPUT = V3 / "gantt-v2.xlsx"

tasks = [
    ("GOV-01", "Proposal handoff", "Tamer (A/R)", "2026-06-08", "2026-06-08", "Done"),
    ("GOV-02", "Requirements and governance baseline", "Tamer (A); Nick and Dan (R)", "2026-06-09", "2026-06-21", "Done"),
    ("GOV-03", "Project coordination through submission", "Tamer (A/R); Dan (R)", "2026-06-09", "2026-08-14", "Doing"),
    ("STR-01", "Market research", "Tamer (A/R); Conrad and Mo (C)", "2026-07-01", "2026-07-07", "Done"),
    ("STR-02", "Strategic analysis", "Tamer (A/R); all leads (C)", "2026-07-08", "2026-07-14", "Done"),
    ("STR-03", "Strategic planning", "Tamer (A/R); all leads (C)", "2026-07-15", "2026-07-21", "Done"),
    ("STR-04", "Financial and legal coordination and oversight", "Tamer (A/R); Conrad and Badhri (C)", "2026-07-01", "2026-07-31", "Doing"),
    ("PRD-01", "Product workflow and release rules", "Dan (A/R); Nick (R)", "2026-06-09", "2026-07-10", "Done"),
    ("PRD-02", "Document rules and pitch alignment", "Dan (A); Nick and Badhri (R); Tamer (C)", "2026-06-22", "2026-08-12", "Doing"),
    ("TEC-01", "Smart contract and contract tests", "Nick (A/R); Dan (R)", "2026-06-15", "2026-07-10", "Done"),
    ("TEC-02", "Application lifecycle, authentication and persistence", "Nick (A/R); Dan (R)", "2026-06-22", "2026-08-05", "Doing"),
    ("TEC-03", "Testnet deployment and onboarding gate", "Nick (A); Dan (R)", "2026-06-30", "2026-08-10", "Doing"),
    ("TEC-04", "Integrated acceptance and clean-build reproducibility", "Nick (A/R); Dan (R)", "2026-08-03", "2026-08-12", "To Do"),
    ("UX-01", "Role journeys and consolidated dashboard", "Mo (A/R); Nick and Dan (C)", "2026-06-22", "2026-08-08", "Doing"),
    ("UX-02", "Canonical demo route and polish", "Mo (A/R); Dan (R); Tamer (C)", "2026-08-03", "2026-08-12", "To Do"),
    ("UX-03", "Front-end UI implementation", "Mo (A/R); Nick and Dan (C); Tamer (I)", "2026-06-22", "2026-08-08", "Doing"),
    ("TEC-05", "Back-end application integration", "Mo (A/R); Nick and Dan (C); Tamer (I)", "2026-07-15", "2026-08-10", "Doing"),
    ("RISK-01", "Legal, regulatory and document research", "Badhri (A/R); Tamer (C)", "2026-06-09", "2026-08-05", "Doing"),
    ("RISK-02", "Final legal and compliance claim review", "Badhri (A/R); Tamer and Nick (C)", "2026-08-03", "2026-08-12", "To Do"),
    ("BUS-01", "Financial model, pricing and funding assumptions", "Conrad (A/R); Mo (R); Tamer (C)", "2026-06-09", "2026-08-12", "Doing"),
    ("BUS-02", "Customer, competitor and market case", "Mo (A/R); Tamer, Conrad and Badhri (C)", "2026-06-09", "2026-08-12", "Doing"),
    ("BUS-03", "Business model and commercial narrative", "Tamer (A/R); Conrad and Badhri (C)", "2026-08-12", "2026-08-12", "To Do"),
    ("SUB-01", "Final deck and script", "Tamer (A/R); Dan and Mo (R); other leads (C)", "2026-07-27", "2026-08-12", "Doing"),
    ("SUB-02", "Demo runbook", "Dan (A/R); Nick (R); Mo (C)", "2026-08-03", "2026-08-12", "To Do"),
    ("SUB-03", "Rehearsals and recordings", "Dan (A/R); Tamer and Mo (R); Nick (C)", "2026-08-10", "2026-08-13", "To Do"),
    ("SUB-04", "Submission QA and upload", "Tamer (A/R); Dan (R); Mo (C)", "2026-08-13", "2026-08-14", "To Do"),
    ("DEF-01", "Post-MVP integrations and production hardening", "Tamer (A); technical and domain leads (C)", None, None, "Deferred"),
]

wb = load_workbook(SOURCE)
ws = wb["Simple Gantt"]
ws.title = "Midway Gantt"

# Extend the styled task area by copying the final source row into two new rows.
for new_row in (34, 35):
    source_row = 33
    ws.row_dimensions[new_row].height = ws.row_dimensions[source_row].height
    for col in range(1, 17):
        src = ws.cell(source_row, col)
        dst = ws.cell(new_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        dst.alignment = copy(src.alignment)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.font = copy(src.font)

ws["A1"] = "Blockmediary — Midway Gantt v2"
ws["A2"] = "Weekly view | 8 June–14 August 2026 | Status date 28 July 2026 | Mid-delivery snapshot"
ws["A6"] = "Bars show the full planned schedule; statuses reflect 28 July 2026.  ◆ Proposal: 8 Jun   ◆ Status: 28 Jul   ◆ Submission: 14 Aug"
ws["O4"] = "ACTIVE PROGRESS"

first_row = 9
last_row = first_row + len(tasks) - 1
for offset, (task_id, task, ownership, start, finish, status) in enumerate(tasks):
    row = first_row + offset
    ws.cell(row, 1, task_id)
    ws.cell(row, 2, task)
    ws.cell(row, 3, ownership)
    ws.cell(row, 4, datetime.fromisoformat(start) if start else None)
    ws.cell(row, 5, datetime.fromisoformat(finish) if finish else None)
    ws.cell(row, 6, status)
    for col in range(7, 17):
        ws.cell(row, col, None)

# KPI formulas and completion formula.
ws["A5"] = f'=COUNTIF($F${first_row}:$F${last_row},"Done")'
ws["D5"] = f'=COUNTIF($F${first_row}:$F${last_row},"Review")'
ws["G5"] = f'=COUNTIF($F${first_row}:$F${last_row},"Doing")'
ws["J5"] = f'=COUNTIF($F${first_row}:$F${last_row},"To Do")'
ws["M5"] = f'=COUNTIF($F${first_row}:$F${last_row},"Deferred")'
ws["O5"] = f'=COUNTIF($F${first_row}:$F${last_row},"Done")/COUNTIF($F${first_row}:$F${last_row},"<>Deferred")'

# Rebuild data validation and formula-driven weekly Gantt bars for all rows.
ws.data_validations.dataValidation = []
dv = DataValidation(type="list", formula1='"Done,Review,Doing,To Do,Deferred"', allow_blank=False)
ws.add_data_validation(dv)
dv.add(f"F{first_row}:F{last_row}")

ws.conditional_formatting._cf_rules.clear()
fills = {
    "Done": "2E7D32",
    "Review": "F9A825",
    "Doing": "1769AA",
    "To Do": "7E57C2",
    "Deferred": "8B95A3",
}
for status, color in fills.items():
    fill = PatternFill("solid", fgColor=color)
    ws.conditional_formatting.add(
        f"G{first_row}:P{last_row}",
        FormulaRule(formula=[f'AND(G$8<=$E{first_row},G$8+6>=$D{first_row},$F{first_row}="{status}")'], fill=fill),
    )
    ws.conditional_formatting.add(
        f"F{first_row}:F{last_row}",
        FormulaRule(formula=[f'$F{first_row}="{status}"'], fill=fill, font=Font(color="FFFFFF", bold=True)),
    )

# Emphasize the week containing the 28 July status date.
status_side = Side(style="medium", color="F9A825")
for row in range(8, last_row + 1):
    cell = ws.cell(row, 14)  # Week beginning 27 July.
    current = cell.border
    cell.border = Border(left=current.left, right=status_side, top=current.top, bottom=current.bottom)
ws["N8"].fill = PatternFill("solid", fgColor="F9A825")
ws["N8"].font = Font(color="FFFFFF", bold=True)

ws.freeze_panes = "G9"
ws.auto_filter.ref = f"A8:P{last_row}"
ws.print_area = f"A1:P{last_row}"
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.print_title_rows = "1:8"
ws.sheet_view.showGridLines = False

# Ensure task rows remain compact but readable.
for row in range(first_row, last_row + 1):
    ws.row_dimensions[row].height = 32
    for col in range(1, 17):
        ws.cell(row, col).alignment = Alignment(
            horizontal="center" if col in (1, 4, 5, 6) or col >= 7 else "left",
            vertical="center", wrap_text=True,
        )

# Clear inherited historical colours and create a correct static preview layer.
# Conditional formatting above remains the live, formula-driven Excel layer.
status_colors = {
    "Done": "2E7D32",
    "Review": "F9A825",
    "Doing": "1769AA",
    "To Do": "7E57C2",
    "Deferred": "8B95A3",
}
base_font = copy(ws["A32"].font)
base_border = copy(ws["A32"].border)
for row in range(first_row, last_row + 1):
    status = ws.cell(row, 6).value
    pale = "FFFFFF" if row % 2 else "F4F7FA"
    for col in range(1, 17):
        cell = ws.cell(row, col)
        cell.font = copy(base_font)
        cell.border = copy(base_border)
        cell.fill = PatternFill("solid", fgColor=pale) if col <= 5 else PatternFill(fill_type=None)
    ws.cell(row, 6).fill = PatternFill("solid", fgColor=status_colors[status])
    ws.cell(row, 6).font = Font(name=base_font.name, size=base_font.size, color="FFFFFF", bold=True)
    start, finish = ws.cell(row, 4).value, ws.cell(row, 5).value
    if start and finish:
        for col in range(7, 17):
            week = ws.cell(8, col).value
            if week <= finish and week + timedelta(days=6) >= start:
                ws.cell(row, col).fill = PatternFill("solid", fgColor=status_colors[status])
    if status == "Deferred":
        for col in range(1, 17):
            ws.cell(row, col).fill = PatternFill("solid", fgColor="ECEFF1")
            ws.cell(row, col).font = Font(name=base_font.name, size=base_font.size,
                                          color="607D8B", italic=True)
        ws.cell(row, 6).fill = PatternFill("solid", fgColor=status_colors[status])
        ws.cell(row, 6).font = Font(name=base_font.name, size=base_font.size,
                                    color="FFFFFF", bold=True)

# Recalculate formulas when opened in Excel.
wb.calculation.fullCalcOnLoad = True
wb.calculation.forceFullCalc = True
wb.calculation.calcMode = "auto"
wb.save(OUTPUT)

# Embed current cached KPI values for renderers that do not calculate formulas.
cached = {"A5": "7", "D5": "0", "G5": "12", "J5": "7", "M5": "1", "O5": str(7 / 26)}
namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
ET.register_namespace("", namespace)
with ZipFile(OUTPUT, "r") as source_zip:
    with NamedTemporaryFile(suffix=".xlsx", delete=False, dir=OUTPUT.parent) as temp_handle:
        temp_path = Path(temp_handle.name)
    with ZipFile(temp_path, "w", ZIP_DEFLATED) as target_zip:
        for item in source_zip.infolist():
            data = source_zip.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                root = ET.fromstring(data)
                for cell in root.findall(f".//{{{namespace}}}c"):
                    ref = cell.attrib.get("r")
                    if ref in cached:
                        value = cell.find(f"{{{namespace}}}v")
                        if value is None:
                            value = ET.SubElement(cell, f"{{{namespace}}}v")
                        value.text = cached[ref]
                data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            target_zip.writestr(item, data)
temp_path.replace(OUTPUT)
print(f"created {OUTPUT}")
