from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(r"C:\Users\cwbec\BlockMed")
WORKBOOK = ROOT / "financial-projections" / "Blockmediary_Financial_Model_Submission_Ready_2026-08-13.xlsx"
OUTPUT = ROOT / "tmp" / "financial-report" / "report-data.json"
YEARS = ["Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
YEAR_COLS = ["C", "D", "E", "F", "G"]


def row_values(ws, row: int, cols: list[str] = YEAR_COLS) -> list[Any]:
    return [ws[f"{col}{row}"].value for col in cols]


def main() -> None:
    values = load_workbook(WORKBOOK, data_only=True, read_only=False)
    formulas = load_workbook(WORKBOOK, data_only=False, read_only=False)

    pnl = values["P&L 5yr"]
    launch = values["Launch Readiness"]
    tier = values["Tier Economics"]
    dashboard = values["Dashboard"]
    cashflow = values["Cash Flow Statement"]
    scenario_engine = values["Scenario Engine"]
    clients = values["Client Funds Memo"]
    checks = values["Model Checks"]
    startup = values["Startup 6mo"]

    operating = {
        "years": YEARS,
        "deals": row_values(pnl, 7),
        "avg_deal_value": row_values(pnl, 8),
        "gmv": row_values(pnl, 9),
        "documents": row_values(pnl, 10),
        "total_revenue": row_values(pnl, 18),
        "cogs": row_values(pnl, 29),
        "gross_profit": row_values(pnl, 30),
        "gross_margin": row_values(pnl, 31),
        "total_fte": row_values(pnl, 37),
        "payroll": row_values(pnl, 42),
        "fixed_opex": row_values(pnl, 56),
        "one_offs": row_values(pnl, 58),
        "ebitda_pre_one_off": row_values(pnl, 61),
        "operating_result": row_values(pnl, 62),
        "cumulative_result": row_values(pnl, 63),
        "peak_deficit": row_values(pnl, 64),
        "reserve": row_values(pnl, 65),
        "funding_need": row_values(pnl, 66),
        "deals_per_active_customer": row_values(pnl, 69),
        "active_customers": row_values(pnl, 70),
        "revenue_per_deal": row_values(pnl, 71),
        "revenue_per_active_customer": row_values(pnl, 72),
        "gross_new_customers": row_values(pnl, 83),
        "new_customers_per_commercial_fte": row_values(pnl, 75),
    }
    operating["revenue_yield"] = [
        revenue / gmv if gmv else 0
        for revenue, gmv in zip(operating["total_revenue"], operating["gmv"])
    ]
    operating["gmv_per_revenue_pound"] = [
        gmv / revenue if revenue else 0
        for revenue, gmv in zip(operating["total_revenue"], operating["gmv"])
    ]

    revenue_breakdown = {
        "years": YEARS,
        "escrow_transaction_fees": row_values(pnl, 14),
        "launch_discounts": row_values(pnl, 15),
        "document_and_exception_fees": row_values(pnl, 16),
        "partner_api_and_ancillary": row_values(pnl, 17),
        "total_revenue": row_values(pnl, 18),
    }

    tier_names = ["Tier A - eBL", "Tier B - carrier/API", "Tier C - paper"]
    tier_cols = ["B", "C", "D"]
    tiers = []
    for name, col in zip(tier_names, tier_cols):
        deals = tier[f"{col}5"].value
        deal_value = tier[f"{col}8"].value
        take_rate = tier[f"{col}10"].value
        minimum = tier[f"{col}11"].value
        review_fee = tier[f"{col}13"].value
        dispute_rate = tier[f"{col}15"].value
        dispute_fee = tier[f"{col}17"].value
        standard_fee = max(deal_value * take_rate, minimum) + review_fee
        expected_revenue_per_deal = tier[f"{col}24"].value / deals
        variable_cost_per_deal = tier[f"{col}23"].value / deals
        tiers.append({
            "name": name,
            "deal_value": deal_value,
            "take_rate": take_rate,
            "minimum_fee": minimum,
            "review_fee": review_fee,
            "dispute_rate": dispute_rate,
            "conditional_dispute_fee": dispute_fee,
            "standard_fee": standard_fee,
            "standard_fee_pct_of_deal": standard_fee / deal_value,
            "deal_value_per_fee_pound": deal_value / standard_fee,
            "expected_revenue_per_deal": expected_revenue_per_deal,
            "variable_cost_per_deal": variable_cost_per_deal,
            "contribution_per_deal_before_shared_cogs": expected_revenue_per_deal - variable_cost_per_deal,
            "tier_gross_margin_before_shared_cogs": tier[f"{col}26"].value,
        })

    tier_mix = {
        "years": YEARS,
        "tier_a": [tier[f"{col}4"].value for col in ["B", "F", "J", "N", "R"]],
        "tier_b": [tier[f"{col}4"].value for col in ["C", "G", "K", "O", "S"]],
        "tier_c": [tier[f"{col}4"].value for col in ["D", "H", "L", "P", "T"]],
        "blended_escrow_take_rate": [tier[f"{col}10"].value for col in ["E", "I", "M", "Q", "U"]],
        "docs_per_deal": [tier[f"{col}6"].value for col in ["E", "I", "M", "Q", "U"]],
        "tier_margin_before_shared_cogs": [tier[f"{col}26"].value for col in ["E", "I", "M", "Q", "U"]],
    }

    scenario_rows = {"Low": (45, "D"), "Base": (46, "J"), "High": (47, "P")}
    scenarios = []
    for name, (row, y3_col) in scenario_rows.items():
        scenarios.append({
            "name": name,
            "year_3_revenue": scenario_engine[f"{y3_col}30"].value,
            "year_3_active_customers": scenario_engine[f"{y3_col}62"].value,
            "year_3_ebitda": scenario_engine[f"{y3_col}57"].value,
            "monthly_break_even": cashflow[f"B{row}"].value,
            "peak_cash_gap": cashflow[f"C{row}"].value,
            "month_60_operating_cash_flow": cashflow[f"D{row}"].value,
            "operating_reserve": cashflow[f"E{row}"].value,
            "funding_need": cashflow[f"F{row}"].value,
        })

    use_of_funds = []
    for row in range(58, 63):
        use_of_funds.append({
            "category": dashboard[f"J{row}"].value,
            "amount": dashboard[f"L{row}"].value,
            "share": dashboard[f"M{row}"].value,
        })

    timing = {
        "first_paid_pilot_month": launch["E6"].value,
        "first_cash_receipt_month": launch["E11"].value,
        "full_vara_licence_month": launch["E14"].value,
        "average_pre_revenue_burn": launch["E13"].value,
        "initial_operating_funding_need": launch["E12"].value,
        "vara_application_fee_gbp": launch["E15"].value,
        "vara_supervision_gbp_per_year": launch["E16"].value,
        "vara_restricted_capital": launch["E17"].value,
        "partner_pilot_one_off": launch["E8"].value,
        "partner_pilot_annual_run_rate": launch["E9"].value,
        "startup_six_month_total_including_contingency": startup["F44"].value,
    }
    timing["combined_staged_capital_envelope"] = (
        timing["initial_operating_funding_need"] + timing["vara_restricted_capital"]
    )

    client_funds = {
        "years": YEARS,
        "holding_period_days": row_values(clients, 7, ["B", "C", "D", "E", "F"]),
        "average_safeguarded_assets": row_values(clients, 8, ["B", "C", "D", "E", "F"]),
        "operating_cash_availability": row_values(clients, 11, ["B", "C", "D", "E", "F"]),
        "recognized_yield": row_values(clients, 12, ["B", "C", "D", "E", "F"]),
    }

    cash_runway = {
        "checkpoints": [dashboard[f"B{row}"].value for row in range(68, 78)],
        "low": [dashboard[f"C{row}"].value * 1_000_000 for row in range(68, 78)],
        "base": [dashboard[f"D{row}"].value * 1_000_000 for row in range(68, 78)],
        "high": [dashboard[f"E{row}"].value * 1_000_000 for row in range(68, 78)],
    }

    reserve_sensitivity = []
    for col in ["C", "D", "E"]:
        reserve_sensitivity.append({
            "months": pnl[f"{col}89"].value,
            "reserve": pnl[f"{col}90"].value,
            "funding_need": pnl[f"{col}91"].value,
        })

    check_statuses = [checks[f"F{row}"].value for row in range(7, 49)]
    formula_errors = []
    for ws in formulas.worksheets:
        value_ws = values[ws.title]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cached = value_ws[cell.coordinate].value
                    if isinstance(cached, str) and cached.startswith("#"):
                        formula_errors.append(f"{ws.title}!{cell.coordinate}: {cached}")

    data = {
        "workbook": str(WORKBOOK),
        "model_date": "2026-08-13",
        "sheet_names": values.sheetnames,
        "operating": operating,
        "revenue_breakdown": revenue_breakdown,
        "tiers": tiers,
        "tier_mix": tier_mix,
        "scenarios": scenarios,
        "use_of_funds": use_of_funds,
        "timing_and_capital": timing,
        "client_funds": client_funds,
        "cash_runway": cash_runway,
        "reserve_sensitivity": reserve_sensitivity,
        "model_checks": {
            "overall_status": checks["B4"].value,
            "count": len(check_statuses),
            "passed": sum(status == "PASS" for status in check_statuses),
            "failed": [
                checks[f"A{row}"].value
                for row in range(7, 49)
                if checks[f"F{row}"].value != "PASS"
            ],
            "formula_errors": formula_errors,
        },
        "source_cells": {
            "five_year_case": "P&L 5yr!C7:G91",
            "tier_economics": "Tier Economics!B3:U35",
            "cash_scenarios": "Cash Flow Statement!B45:F47",
            "launch_readiness": "Launch Readiness!A5:J23",
            "use_of_funds": "Dashboard!J57:M62",
            "client_funds": "Client Funds Memo!A5:H31",
            "checks": "Model Checks!A1:G48",
        },
    }

    OUTPUT.write_text(json.dumps(data, indent=2), encoding="utf-8")
    print(json.dumps({
        "overall_status": data["model_checks"]["overall_status"],
        "checks": f"{data['model_checks']['passed']}/{data['model_checks']['count']}",
        "formula_errors": len(formula_errors),
        "initial_funding": timing["initial_operating_funding_need"],
        "year_3_revenue": operating["total_revenue"][2],
        "break_even": next(s["monthly_break_even"] for s in scenarios if s["name"] == "Base"),
    }, indent=2))


if __name__ == "__main__":
    main()
